# -*- coding: utf-8 -*-
'''
Created on 22.02.2021

@author: Luca_
'''

import tkinter as tk
import tkinter.ttk as TTK
from openpyxl import load_workbook

from _datetime import datetime
from tkcalendar import DateEntry


# Sheet zur Bestimmung DURCHSCHNITT der ENDPLATZIERUNGEN *nach* ENTLASSUNG 
alleGefeuertenTrainer = load_workbook(filename="excelData/alleGefeuertenTrainer.xlsx")  # should be alleGefezertenTrainer.xlsx
sheetGefeuerteTrainer = alleGefeuertenTrainer.active

# Sheet zur Bestimmung DURCHSCHNITT der END PPS *nach* ENTLASSUNG 
saisonTabellenOhneGefeuerteTrainerC = load_workbook(filename="excelData/saisonTabelleOhneGefeuerteTrainerC.xlsx")
sheetSaisonTabellenOhneGefeuerteTrainerC = saisonTabellenOhneGefeuerteTrainerC.active

# Sheet zur Bestimmung DURCHSCHNITT der ENDPLATZIERUNGEN und END PPS *vor* ENTLASSUNG 
saisonTabellenOhneGefeuerteTrainerD = load_workbook(filename="excelData/saisonTabelleOhneGefeuerteTrainerD.xlsx")
sheetSaisonTabellenOhneGefeuerteTrainerD = saisonTabellenOhneGefeuerteTrainerD.active


########################################################################
clubsSaisonsGefeuert = {}
spielTageundRänge = {}

printAktivierer = 0

def berechneDurchschnittlichePlatzierungenZumEntlassungsZeitpunktUndEndPlatzierungenBeiGefeuertenTrainern(spielT, lSpielT, platzierung, Eplatzierung, datumswert):
    x = 2
    spielTagsWert = int(spielT)
    endSpieltagsWert = int(lSpielT)
    platzierungsWert = int(platzierung)
    endPlatzierungsWert = int(Eplatzierung)
    datumsWert = datumswert
    dt = datetime.combine(datumsWert, datetime.min.time())
    # dt = datetime.strptime(datumsWert, '%d/%m/%Y')
    
    allePlatzierungenZusammen = 0
    alleEndPlatzierungenZusammen = 0
    anzahlDerGezähltenPlatzierungen = 0

    global clubsSaisonsGefeuert
    global spielTageundRänge

    for _ in sheetGefeuerteTrainer.iter_rows(min_row=2, max_row=len(sheetGefeuerteTrainer["A"]), max_col=7, values_only=True):
        try:
            if (int(sheetGefeuerteTrainer["F" + str(x)].value) >= spielTagsWert) and (int(sheetGefeuerteTrainer["F" + str(x)].value < endSpieltagsWert)) and (int(sheetGefeuerteTrainer["B" + str(x)].value >= platzierungsWert)) and (int(sheetGefeuerteTrainer["B" + str(x)].value <= endPlatzierungsWert)) and (sheetGefeuerteTrainer["E" + str(x)].value >= dt) and (str(sheetGefeuerteTrainer["E" + str(x)].value.year) < "2021"):
                
                club = str(sheetGefeuerteTrainer["A" + str(x)].value)
                meineSaison = str(sheetGefeuerteTrainer["D" + str(x)].value)
                
                if club not in clubsSaisonsGefeuert: 
                    clubsSaisonsGefeuert[club] = []
                    
                if meineSaison not in clubsSaisonsGefeuert[club]:
                    clubsSaisonsGefeuert[club].append(meineSaison)
                    allePlatzierungenZusammen += int(sheetGefeuerteTrainer["B" + str(x)].value)
                    alleEndPlatzierungenZusammen += int(sheetGefeuerteTrainer["C" + str(x)].value)
                    anzahlDerGezähltenPlatzierungen += 1
                    
                spielTagsEintrag = str(sheetGefeuerteTrainer["F" + str(x)].value)
                platzierungsEintrag = str(sheetGefeuerteTrainer["B" + str(x)].value)
                
                if spielTagsEintrag not in spielTageundRänge: 
                    spielTageundRänge[spielTagsEintrag] = []
                if platzierungsEintrag not in spielTageundRänge[spielTagsEintrag]:
                    spielTageundRänge[spielTagsEintrag].append(platzierungsEintrag)
                
        except:
            print ("Error in berechneDurchschnittlichePlatzierungenZumEntlassungsZeitpunktUndEndPlatzierungenBeiGefeuertenTrainern")
        x += 1
    global printAktivierer
    if(printAktivierer == 0): 
        print("Clubs, die ihren Trainer gefeuert haben, sowie die jeweilige Saison: " + str(clubsSaisonsGefeuert))
        print("Spieltage an denen Trainer gefeuert wurden, sowie der jeweilige Rang: " + str(spielTageundRänge))
        print("Alle Platzierungen dieser Mannschaften- zum Entlassungszeitpunkt des Trainers, zusammengezählt: " + str(allePlatzierungenZusammen))
        print("Alle Platzierungen dieser Mannschaften am Ende der Saison " + str(alleEndPlatzierungenZusammen))
        print("Anzahl der gezählten Mannschaften: " + str(anzahlDerGezähltenPlatzierungen))
        print("")
        printAktivierer = 1
    else:
        printAktivierer = 0
        
    return [str(allePlatzierungenZusammen / anzahlDerGezähltenPlatzierungen),str(alleEndPlatzierungenZusammen/anzahlDerGezähltenPlatzierungen)]



def berechneDurchschnittlichePlatzierungenUndEndPlatzierungenBeiNichtGefeuertenTrainern(datumswert):
    x = 2
    datumsWert = datumswert
    dt = datetime.combine(datumsWert, datetime.min.time())
    
    allePlatzierungenZusammen = 0
    alleEndPlatzierungenZusammen = 0
    anzahlDerGezähltenPlatzierungen = 0
    
    clubsDatums1 = { }
    
    global clubsSaisonsGefeuert
    global spielTageundRänge

    for _ in sheetSaisonTabellenOhneGefeuerteTrainerD.iter_rows(min_row=2, max_row=len(sheetSaisonTabellenOhneGefeuerteTrainerD["A"]), max_col=8, values_only=True):
        try:
            if (str(sheetSaisonTabellenOhneGefeuerteTrainerD["D" + str(x)].value) in spielTageundRänge) and (str(sheetSaisonTabellenOhneGefeuerteTrainerD["B" + str(x)].value) in spielTageundRänge[str(sheetSaisonTabellenOhneGefeuerteTrainerD["D" + str(x)].value)]) and (sheetSaisonTabellenOhneGefeuerteTrainerD["F" + str(x)].value >= dt) and (str(sheetSaisonTabellenOhneGefeuerteTrainerC["F" + str(x)].value.year) < "2021"):
                if ((str(sheetSaisonTabellenOhneGefeuerteTrainerD["A" + str(x)].value) in clubsSaisonsGefeuert)):
                    if(str(sheetSaisonTabellenOhneGefeuerteTrainerD["E" + str(x)].value) in clubsSaisonsGefeuert[str(sheetSaisonTabellenOhneGefeuerteTrainerD["A" + str(x)].value)]):
                        x += 1 
                        continue
                
                allePlatzierungenZusammen += int(sheetSaisonTabellenOhneGefeuerteTrainerD["B" + str(x)].value)
                alleEndPlatzierungenZusammen += int(sheetSaisonTabellenOhneGefeuerteTrainerD["C" + str(x)].value)
                
                club = str(sheetSaisonTabellenOhneGefeuerteTrainerD["A" + str(x)].value)
                meineSaison = str(sheetSaisonTabellenOhneGefeuerteTrainerD["E" + str(x)].value)
                
                if club not in clubsDatums1: 
                    clubsDatums1[club] = {}
                
                if(meineSaison not in clubsDatums1[club]):
                    clubsDatums1[club][meineSaison] = []
                    
                clubsDatums1[club][meineSaison].append(1)
                
                anzahlDerGezähltenPlatzierungen += 1
                
        except:
            print ("Error in berechneDurchschnittlichePlatzierungenUndEndPlatzierungenBeiNichtGefeuertenTrainern")
        x += 1


    global printAktivierer
    if(printAktivierer == 0): 
        print("Clubs, die ihren Trainer NICHT gefeuert haben, sowie die jeweilige Saison: " + str(clubsDatums1))
        print("Alle Platzierungen dieser Mannschaften zusammengezählt: " + str(allePlatzierungenZusammen))
        print("Alle Platzierungen dieser Mannschaften am Ende der Saison " + str(alleEndPlatzierungenZusammen))
        print("Anzahl der gezählten Mannschaften: " + str(anzahlDerGezähltenPlatzierungen))
        print("")
        printAktivierer = 1
    else:
        printAktivierer = 0
    return [str(allePlatzierungenZusammen / anzahlDerGezähltenPlatzierungen),str(alleEndPlatzierungenZusammen/anzahlDerGezähltenPlatzierungen)]

    
def berechneDurchschnittlichePPS_BeiGefeuertenTrainern(spielT, letzterSpieltag, platzierung, maxPlatzierung, datumswert):
    x = 2
    
    finalerSpieltag = 34
    
    spielTagsWert = int(spielT)
    letzterSpieltagsWert = int(letzterSpieltag)
    platzierungsWert = int(platzierung)
    maxPlatzierungsWert = int(maxPlatzierung)
    datumsWert = datumswert
    _1963und1964datumsChecker = datetime(1965, 6, 1)
    _1991datumsChecker = "1991/1992"
    gespeicherterLetzterSpieltagsWert = letzterSpieltagsWert
    dt = datetime.combine(datumsWert, datetime.min.time())
    # dt = datetime.strptime(datumsWert, '%d/%m/%Y')

    allePunkteZuBeginn = 0
    anzahlAnGezähltenInstanzen = 0
    
    allePunkteZumSchluss = 0
    
    
    global clubsSaisonsGefeuert
    anzahlSpieleListe = []
    ppsBisZuDiesemZeitpunktListe = []
    
    for _ in sheetGefeuerteTrainer.iter_rows(min_row=2, max_row=len(sheetGefeuerteTrainer["A"]), max_col=7, values_only=True):
        if (sheetGefeuerteTrainer["E" + str(x)].value <= _1963und1964datumsChecker):
            finalerSpieltag = 30
            if (letzterSpieltagsWert > 30):
                letzterSpieltagsWert = 30
        elif (sheetGefeuerteTrainer["D" + str(x)].value == _1991datumsChecker):
            letzterSpieltagsWert = gespeicherterLetzterSpieltagsWert
            finalerSpieltag =38
        else:
            if (letzterSpieltagsWert>34):
                letzterSpieltagsWert = 34
            else:
                letzterSpieltagsWert = gespeicherterLetzterSpieltagsWert
            finalerSpieltag = 34

        try:
            if (int(sheetGefeuerteTrainer["F" + str(x)].value) >= spielTagsWert) and (int(sheetGefeuerteTrainer["F" + str(x)].value) < letzterSpieltagsWert) and (int(sheetGefeuerteTrainer["B" + str(x)].value) >= platzierungsWert) and (int(sheetGefeuerteTrainer["B" + str(x)].value) <= maxPlatzierungsWert) and (sheetGefeuerteTrainer["E" + str(x)].value >= dt and (str(sheetGefeuerteTrainer["E" + str(x)].value.year) < "2021")):
                allePunkteZuBeginn += int(sheetGefeuerteTrainer["G" + str(x)].value)

                club = str(sheetGefeuerteTrainer["A" + str(x)].value)
                meineSaison = str(sheetGefeuerteTrainer["D" + str(x)].value)
                
                if club not in clubsSaisonsGefeuert: 
                    clubsSaisonsGefeuert[club] = []
                if meineSaison not in clubsSaisonsGefeuert[club]:
                    clubsSaisonsGefeuert[club].append(meineSaison)

                    
                spielTagsEintrag = str(sheetGefeuerteTrainer["F" + str(x)].value)
                platzierungsEintrag = str(sheetGefeuerteTrainer["B" + str(x)].value)
                
                if spielTagsEintrag not in spielTageundRänge: 
                    spielTageundRänge[spielTagsEintrag] = []
                if platzierungsEintrag not in spielTageundRänge[spielTagsEintrag]:
                    spielTageundRänge[spielTagsEintrag].append(platzierungsEintrag)

                anzahlSpieleListe.append(letzterSpieltagsWert - sheetGefeuerteTrainer["F"+str(x)].value)
                ppsBisZuDiesemZeitpunktListe.append(float(sheetGefeuerteTrainer["G" + str(x)].value/float(sheetGefeuerteTrainer["F" + str(x)].value)))
        except:
            print("Error in berechneDurchschnittlichePPS_BeiGefeuertenTrainern")
        x += 1
        
    ##pps seitdem
    x = 2
    for _ in sheetSaisonTabellenOhneGefeuerteTrainerC.iter_rows(min_row=2, max_row=len(sheetSaisonTabellenOhneGefeuerteTrainerC["A"]), max_col=8, values_only=True):
            if(sheetSaisonTabellenOhneGefeuerteTrainerC["D" + str(x)].value == finalerSpieltag) and (str(sheetSaisonTabellenOhneGefeuerteTrainerC["A" + str(x)].value) in clubsSaisonsGefeuert) and (str(sheetSaisonTabellenOhneGefeuerteTrainerC["E" + str(x)].value) in clubsSaisonsGefeuert[str(sheetSaisonTabellenOhneGefeuerteTrainerC["A" + str(x)].value)]):
                allePunkteZumSchluss += int(sheetSaisonTabellenOhneGefeuerteTrainerC["G" + str(x)].value)
                anzahlAnGezähltenInstanzen += 1
                
            x += 1
    
    durchschnittlicheAnzahlSpiele = sum(anzahlSpieleListe) / len(anzahlSpieleListe)
    gesamtPPS_BisZuDiesemZeitpunkt = str(sum(ppsBisZuDiesemZeitpunktListe)/len(ppsBisZuDiesemZeitpunktListe))
    gesamtPPS_AbDiesemZeitpunkt = str(((allePunkteZumSchluss - allePunkteZuBeginn) / anzahlAnGezähltenInstanzen) / (durchschnittlicheAnzahlSpiele))
    global printAktivierer
    if(printAktivierer == 0): 
        print("Clubs, die ihren Trainer gefeuert haben, sowie die jeweilige Saison: " + str(clubsSaisonsGefeuert))
        print("Alle Punkte dieser Mannschaften- zum Entlassungszeitpunkt des Trainers, zusammengezählt: " + str(allePunkteZuBeginn))
        print("Alle Punkte dieser Mannschaften am Ende der Saison " + str(allePunkteZumSchluss))
        print("Anzahl der gezählten Mannschaften: " + str(anzahlAnGezähltenInstanzen))
        print("Durchschnittliche PPS bis zu diesem Zeitpunkt: " + gesamtPPS_BisZuDiesemZeitpunkt)
        print("Durchschnittliche Anzahl an Spielen: " + str(durchschnittlicheAnzahlSpiele))
        print("")
        printAktivierer = 1
    else:
        printAktivierer = 0
    return [gesamtPPS_BisZuDiesemZeitpunkt,gesamtPPS_AbDiesemZeitpunkt ]


def berechneDurchschnittlichePPS_BeiNichtGefeuertenTrainern(spielT, letzterSpieltag, platzierung, maxPlatzierung, datumswert):
    global clubsSaisonsGefeuert
    x = 2
    letzterSpieltagsWert = int(letzterSpieltag)
    gespeicherterLetzterSpieltagsWert = letzterSpieltagsWert
    datumsWert = datumswert
    dt = datetime.combine(datumsWert, datetime.min.time())
    # dt = datetime.strptime(datumsWert, '%d/%m/%Y')

    allePunkteZumBeginn = 0
    anzahlAnGezähltenInstanzen = 0
    
    allePunkteZumSchluss = 0
    
    clubsDatums1 = { }

    
    spielTagsListe = []
    
    ppsBisZuDiesemZeitpunktListe = []
    
    _1963und1964datumsChecker = datetime(1965, 6, 1)
    _1991datumsChecker = "1991/1992"
        
    
    for _ in sheetSaisonTabellenOhneGefeuerteTrainerD.iter_rows(min_row=2, max_row=len(sheetSaisonTabellenOhneGefeuerteTrainerD["A"]), max_col=8, values_only=True):
        if (sheetSaisonTabellenOhneGefeuerteTrainerD["F" + str(x)].value <= _1963und1964datumsChecker):
            finalerSpieltag = 30
            if (letzterSpieltagsWert > 30):
                letzterSpieltagsWert = 30
        elif (sheetSaisonTabellenOhneGefeuerteTrainerD["E" + str(x)].value == _1991datumsChecker):
            letzterSpieltagsWert = gespeicherterLetzterSpieltagsWert
            finalerSpieltag =38
        else:
            if (letzterSpieltagsWert>34):
                letzterSpieltagsWert = 34
            else:
                letzterSpieltagsWert = gespeicherterLetzterSpieltagsWert
            finalerSpieltag = 34
        try:
            if (str(sheetSaisonTabellenOhneGefeuerteTrainerD["D" + str(x)].value) in spielTageundRänge) and (str(sheetSaisonTabellenOhneGefeuerteTrainerD["B" + str(x)].value) in spielTageundRänge[str(sheetSaisonTabellenOhneGefeuerteTrainerD["D" + str(x)].value)]) and (sheetSaisonTabellenOhneGefeuerteTrainerD["F" + str(x)].value >= dt) and (str(sheetSaisonTabellenOhneGefeuerteTrainerD["F" + str(x)].value.year) < "2021"):
                # allePunkteZumBeginn += int(sheetSaisonTabellenOhneGefeuerteTrainerD["G"+str(x)].value)
                # HIER FEHLER -> weil wir punkte auch zählen, obwohl trainer am @letzterSpielTag gefeuert werden könnte
                # print(row)
                if ((str(sheetSaisonTabellenOhneGefeuerteTrainerD["A" + str(x)].value) in clubsSaisonsGefeuert)):
                    if(str(sheetSaisonTabellenOhneGefeuerteTrainerD["E" + str(x)].value) in clubsSaisonsGefeuert[str(sheetSaisonTabellenOhneGefeuerteTrainerD["A" + str(x)].value)]):
                        x += 1 
                        continue
                club = str(sheetSaisonTabellenOhneGefeuerteTrainerD["A" + str(x)].value)
                meineSaison = str(sheetSaisonTabellenOhneGefeuerteTrainerD["E" + str(x)].value)
                
                if club not in clubsDatums1: 
                    clubsDatums1[club] = {}
                
                if(meineSaison not in clubsDatums1[club]):
                    clubsDatums1[club][meineSaison] = []
                    
                clubsDatums1[club][meineSaison].append(1)
                
                anzahlAnGezähltenInstanzen += 1
                allePunkteZumBeginn += int(sheetSaisonTabellenOhneGefeuerteTrainerD["G" + str(x)].value)
                
                spielTagsListe.append(letzterSpieltagsWert - sheetSaisonTabellenOhneGefeuerteTrainerD["D"+str(x)].value)
                ppsBisZuDiesemZeitpunktListe.append(float(sheetSaisonTabellenOhneGefeuerteTrainerD["G" + str(x)].value/float(sheetSaisonTabellenOhneGefeuerteTrainerD["D" + str(x)].value)))

        except:
            print ("Error 1 in berechneDurchschnittlichePPS_BeiNichtGefeuertenTrainern")
        x += 1
        
    x = 2
    
    for _ in sheetSaisonTabellenOhneGefeuerteTrainerD.iter_rows(min_row=2, max_row=len(sheetSaisonTabellenOhneGefeuerteTrainerD["A"]), max_col=8, values_only=True):
        try:
            if(sheetSaisonTabellenOhneGefeuerteTrainerD["D" + str(x)].value == finalerSpieltag) and (str(sheetSaisonTabellenOhneGefeuerteTrainerD["A" + str(x)].value) in clubsDatums1) and (str(sheetSaisonTabellenOhneGefeuerteTrainerD["E" + str(x)].value) in clubsDatums1[str(sheetSaisonTabellenOhneGefeuerteTrainerD["A" + str(x)].value)]):
                anzahlVorkommnisseInnerhalbEinesJahres = len(clubsDatums1[str(sheetSaisonTabellenOhneGefeuerteTrainerD["A" + str(x)].value)][sheetSaisonTabellenOhneGefeuerteTrainerD["E" + str(x)].value])
                allePunkteZumSchluss += int(sheetSaisonTabellenOhneGefeuerteTrainerD["G" + str(x)].value)*anzahlVorkommnisseInnerhalbEinesJahres
                
        except:
            print("Error 2 in berechneDurchschnittlichePPS_BeiNichtGefeuertenTrainern")
        x += 1
      
    durchschnittlicheAnzahlSpiele = sum(spielTagsListe) / len(spielTagsListe)
    #print(spielTagsListe)
    gesamtPPS_BisZuDiesemZeitpunkt = str(sum(ppsBisZuDiesemZeitpunktListe)/len(ppsBisZuDiesemZeitpunktListe))
    gesamtPPS_AbDiesemZeitpunkt = str(((allePunkteZumSchluss - allePunkteZumBeginn) / anzahlAnGezähltenInstanzen) / (durchschnittlicheAnzahlSpiele))

    
    global printAktivierer
    if(printAktivierer == 0): 
        print("Clubs, die ihren Trainer NICHT gefeuert haben, sowie die jeweilige Saison: " + str(clubsDatums1))
        print("Alle Punkte dieser Mannschaften zusammengezählt: " + str(allePunkteZumBeginn))
        print("Alle Punkte dieser Mannschaften am Ende der Saison " + str(allePunkteZumSchluss))
        print("Anzahl der gezählten Mannschaften: " + str(anzahlAnGezähltenInstanzen))
        print("Durchschnittliche PPS bis zu diesem Zeitpunkt: " + gesamtPPS_BisZuDiesemZeitpunkt)
        print("Durchschnittliche Anzahl an Spielen: " + str(durchschnittlicheAnzahlSpiele))
        print("")
        printAktivierer = 1
    else:
        printAktivierer = 0
    return [gesamtPPS_BisZuDiesemZeitpunkt,gesamtPPS_AbDiesemZeitpunkt]


def clearLists():
    clubsSaisonsGefeuert.clear()
    spielTageundRänge.clear()

def updateCombobox(cb, data): 
    # put new data 
    cb['values'] = data 


def handle_click():
    tbResults.delete("1.0", tk.END)
        
    if(int(cbSpieltage.get()) >= int(cbLetzteSpieltage.get())):
        print("Anfangs Spieltag darf nicht größer sein als End Spieltag")
        return
    if(int(cbTabellenplätze.get()) > int(cbMaxTabellenplätze.get())):
        print("Anfangs Tabellenplatzauswahl darf nicht größer sein als End Tabellenplatzauswahl")
        return
    if(int(cbSpieltage.get()) in spieltage):
        if(int(cbTabellenplätze.get()) in tabellenplätze):
                
            if(platzierungsVar.get() == 1):
                ergebnisEntlassungsPlatzierung = berechneDurchschnittlichePlatzierungenZumEntlassungsZeitpunktUndEndPlatzierungenBeiGefeuertenTrainern(cbSpieltage.get(), cbLetzteSpieltage.get(), cbTabellenplätze.get(), cbMaxTabellenplätze.get(), calDatum.get_date())[0]
                clearLists()
                ergebnisEndPlatzierung = berechneDurchschnittlichePlatzierungenZumEntlassungsZeitpunktUndEndPlatzierungenBeiGefeuertenTrainern(cbSpieltage.get(), cbLetzteSpieltage.get(), cbTabellenplätze.get(), cbMaxTabellenplätze.get(), calDatum.get_date())[1]
                ergebnisNichtGefeuertePlatzierung = berechneDurchschnittlichePlatzierungenUndEndPlatzierungenBeiNichtGefeuertenTrainern(calDatum.get_date())[0]
                ergebnisNichtGefeuertEndplatzierung = berechneDurchschnittlichePlatzierungenUndEndPlatzierungenBeiNichtGefeuertenTrainern(calDatum.get_date())[1]
                clearLists()
            if(ppsVar.get() == 1):
                ppsNachEntlassung = berechneDurchschnittlichePPS_BeiGefeuertenTrainern(cbSpieltage.get(), cbLetzteSpieltage.get() , cbTabellenplätze.get(), cbMaxTabellenplätze.get(), calDatum.get_date())[1]
                ppsVorEntlassung = berechneDurchschnittlichePPS_BeiGefeuertenTrainern(cbSpieltage.get(), cbLetzteSpieltage.get() , cbTabellenplätze.get(), cbMaxTabellenplätze.get(), calDatum.get_date())[0]
                ppsBeiFesthaltenAnTrainer = berechneDurchschnittlichePPS_BeiNichtGefeuertenTrainern(cbSpieltage.get(), cbLetzteSpieltage.get() , cbTabellenplätze.get(), cbMaxTabellenplätze.get(), calDatum.get_date())[1]
                ppsBeiFesthaltenAnTrainerVorher = berechneDurchschnittlichePPS_BeiNichtGefeuertenTrainern(cbSpieltage.get(), cbLetzteSpieltage.get() , cbTabellenplätze.get(), cbMaxTabellenplätze.get(), calDatum.get_date())[0]
                clearLists()
                
            tbResults.insert(tk.END, "AUSGEWÄHLTE PARAMETER: ")
            tbResults.insert(tk.END, "\nDatum: " + str(calDatum.get_date()))
            #tbResults.insert("3.0", "\n" + cbMannschaften.get())
            tbResults.insert(tk.END, "\nAb Spieltag: " + str(cbSpieltage.get()))
            tbResults.insert(tk.END, "\nBis Spieltag: " + str(cbLetzteSpieltage.get()))
            tbResults.insert(tk.END, "\nAb Tabellenplatz " + str(cbTabellenplätze.get()))
            tbResults.insert(tk.END, "\nBis Tabellenplatz " + str(cbMaxTabellenplätze.get()))
            tbResults.insert(tk.END, "\n")
            try:
                if(platzierungsVar.get() == 1):
                    tbResults.insert(tk.END, "\nDurchschnittliche Platzierung von Mannschaften, zum Entlassungszeitpunkt \ndes Trainers: \n" + ergebnisEntlassungsPlatzierung)
                    tbResults.insert(tk.END, "\nDurchschnittliche Platzierungen von Mannschaften, die in ähnlichen Situationen \nden Trainer NICHT feuerten: \n" + ergebnisNichtGefeuertePlatzierung)
                
                    tbResults.insert(tk.END, "\n")
                
                    tbResults.insert(tk.END, "\nDurchschnittliche Endplatzierung von Mannschaften,\ndie den Trainer feuerten: \n" + ergebnisEndPlatzierung)
                    tbResults.insert(tk.END, "\nDurchschnittliche Endplatzierung von Mannschaften, die in ähnlichen Situationen \nden Trainer NICHT feuerten: \n" + ergebnisNichtGefeuertEndplatzierung)

                    tbResults.insert(tk.END, "\n")
                if(ppsVar.get() == 1):
                    tbResults.insert(tk.END, "\nDurchschnittliche PPS von Mannschaften,\ndie den Trainer feuerten: \n" + ppsNachEntlassung)
                    tbResults.insert(tk.END, "\nDurchschnittliche PPS von diesen Mannschaften,\nbevor sie den Trainer feuerten: \n" + ppsVorEntlassung)
                    
                    tbResults.insert(tk.END, "\n")
                    
                    tbResults.insert(tk.END, "\nDurchschnittliche PPS von Mannschaften, die in ähnlichen Situationen \nden Trainer NICHT feuerten: \n" + ppsBeiFesthaltenAnTrainer)
                    tbResults.insert(tk.END, "\nDurchschnittliche PPS von diesen Mannschaften, \nbis zu dem Zeitpunkt an dem die entlassung hätte stattfinden können: \n" + ppsBeiFesthaltenAnTrainerVorher)
                    
                    tbResults.insert(tk.END, "\n")     
            except:
                print("Keine Gültigen Parameter ausgewählt")
    else:
        print("Keine Gültigen Parameter ausgewählt")

##################################################################


window = tk.Tk()

# Creating Greeting Label
greeting = tk.Message(width=500,text="Berechne die durchschnittlichen Platzierungen und Punkte Pro Spiel (PPS) aller Mannschaften, die -beginnend mit ausgewählten Parametern, ihren Trainer gefeuert- bzw. nicht gefeuert haben")
greeting.pack()

# Creating datums label
datumsLabel = tk.Label(text="Datum auswählen: ")
datumsLabel.pack()

# creating Calendar
calDatum = DateEntry(window, width=30, bg="darkblue", fg="white", date_pattern='mm/dd/y', year=2010, month=8, day=1)
# calDatum = Calendar()
calDatum.pack()


# Creating Spielagsauswahl LAbel
spieltagsLabel = tk.Label(text="Von Spieltag Nr: ")
spieltagsLabel.pack()

# Creating Spieltags Combobox
cbSpieltage = TTK.Combobox()
cbSpieltage.pack()
spieltage = (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40)
cbSpieltage.set(20)
updateCombobox(cbSpieltage, spieltage)

# Createing letzter Spieltags Label
letzterSpieltagsLabel = tk.Label(text="Bis Spieltag Nr:")
letzterSpieltagsLabel.pack()

# Creating letzter Spieltags Combobox
cbLetzteSpieltage = TTK.Combobox()
cbLetzteSpieltage.pack()
letzteSpieltage = (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38)
cbLetzteSpieltage.set(34)
updateCombobox(cbLetzteSpieltage, spieltage)

# Creating Tabellenplatz  label
tabellenLabel = tk.Label(text="Tabellenplatz auswählen: ")
tabellenLabel.pack()

# Creating Tabellenplatz Combobox
cbTabellenplätze = TTK.Combobox()
cbTabellenplätze.pack()
tabellenplätze = (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20)
cbTabellenplätze.set(15)
updateCombobox(cbTabellenplätze, tabellenplätze)

# Creating maxTabellenplatz  label
maxTabellenlabel = tk.Label(text="Maximalen Tabellenplatz auswählen: ")
maxTabellenlabel.pack()

# Creating maxTabellenplatz Combobox
cbMaxTabellenplätze = TTK.Combobox()
cbMaxTabellenplätze.pack()
cbMaxTabellenplätze.set(18)
updateCombobox(cbMaxTabellenplätze, tabellenplätze)

# Creating Checkbuttons for pps and finale platzierungen
ppsVar = tk.IntVar()
platzierungsVar = tk.IntVar()
cButtonPPS = tk.Checkbutton(text="PPS", variable=ppsVar)
cButtonPPS.pack()
cButtonPlatzierungen = tk.Checkbutton(text="Tabellenplätze", variable=platzierungsVar)
cButtonPlatzierungen.pack()

# Creating Confirm Button
button = tk.Button(
    text="Bestätige!",
    width=10,
    height=2,
    command=lambda:[handle_click()]
    )
button.pack()


# Creating Text Box to display results
tbResults = tk.Text()
tbResults.pack()

window.mainloop()



